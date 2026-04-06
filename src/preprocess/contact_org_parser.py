from __future__ import annotations

"""
contact_org_parser.py
---------------------
연락망/조직표형 Excel을 파싱해서 조직 계층 + 인물 정보를 표준 스키마로 반환한다.

핵심 아이디어
- 병합셀 값을 먼저 복원한다.
- 빈칸을 상하/좌우 forward fill 해서 상위 조직명을 살린다.
- 헤더 후보(부서/성명/직책/내선/휴대폰/e-mail)를 탐지한다.
- 좌우 2단 배치 같은 wide sheet는 "블록"으로 분리해서 각 블록을 독립 파싱한다.
- 최종 컬럼을 org_level_1, org_level_2, org_level_3, org, dept, name, title, phone, email 로 표준화한다.

주의
- 연락망 양식은 회사마다 편차가 크므로 100% 완전 자동보다 "안전한 추출 + 검토 가능"을 목표로 한다.
- 병합셀/안내문/합계행이 많아도 최대한 조직 단서를 보존하도록 설계했다.
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 공통 시그널
# ---------------------------------------------------------------------------
NAME_KEYS = ["성명", "이름", "사원명", "직원명", "name"]
DEPT_KEYS = ["부서", "팀", "조직", "소속", "사업부", "사업본부", "본부"]
TITLE_KEYS = ["직책", "직급", "직위", "호칭", "title"]
PHONE_KEYS = ["휴대폰", "핸드폰", "전화", "연락처", "mobile", "phone"]
EMAIL_KEYS = ["e-mail", "email", "메일", "이메일"]
EXT_KEYS = ["내선", "내선번호", "extension"]
ORG_LIKE_KEYS = ["본부", "사업부", "사업본부", "센터", "실", "그룹", "랩", "lab", "팀", "파트"]
SKIP_TEXT = ["총원", "합계", "설명", "비고", "직원연락망", "연락망"]


@dataclass
class ParseBlockResult:
    sheet_name: str
    block_index: int
    rows: pd.DataFrame


def _clean_text(x: object) -> str:
    if x is None:
        return ""
    s = str(x).replace("\n", " ").replace("\r", " ").strip()
    return " ".join(s.split())


def _norm_key(x: object) -> str:
    return _clean_text(x).lower().replace(" ", "")


def _contains_any(text: object, keys: Iterable[str]) -> bool:
    t = _norm_key(text)
    if not t:
        return False
    return any(_norm_key(k) in t or t in _norm_key(k) for k in keys if _norm_key(k))


def _is_empty_val(x: object) -> bool:
    t = _clean_text(x)
    return t in {"", "nan", "none", "<na>"}


def _looks_like_person_name(x: object) -> bool:
    t = _clean_text(x)
    if not t:
        return False

    # 🔥 길이 제한 완화
    if len(t) > 30:
        return False

    # 숫자 너무 많으면 제외
    if sum(c.isdigit() for c in t) > 3:
        return False
    
    if _contains_any(t, SKIP_TEXT + DEPT_KEYS + TITLE_KEYS + PHONE_KEYS + EMAIL_KEYS + EXT_KEYS):
        return False
    return True


def _looks_like_email(x: object) -> bool:
    t = _clean_text(x)
    return "@" in t and "." in t


def _looks_like_phone(x: object) -> bool:
    t = _clean_text(x).replace("-", "").replace(" ", "")
    digits = "".join(ch for ch in t if ch.isdigit())
    return len(digits) >= 7


def _looks_like_org_name(x: object) -> bool:
    t = _clean_text(x)
    if not t or _looks_like_person_name(t):
        return False
    return _contains_any(t, ORG_LIKE_KEYS)


# ---------------------------------------------------------------------------
# Excel 읽기: 병합셀 복원 + wide sheet를 pandas로 변환
# ---------------------------------------------------------------------------
def _worksheet_to_df_merged_values(ws) -> pd.DataFrame:
    """병합 영역 값을 모두 복원한 뒤 DataFrame으로 변환."""
    data = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    if not data:
        return pd.DataFrame()

    # 1-indexed 보정
    for merged in list(ws.merged_cells.ranges):
        min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        val = data[min_row - 1][min_col - 1]
        for r in range(min_row - 1, max_row):
            for c in range(min_col - 1, max_col):
                if data[r][c] is None:
                    data[r][c] = val

    # 열 길이 보정
    width = max(len(r) for r in data)
    data = [list(r) + [None] * (width - len(r)) for r in data]
    return pd.DataFrame(data)


def _split_wide_blocks(raw: pd.DataFrame, min_nonempty_cols: int = 2) -> List[pd.DataFrame]:
    """좌우 2단/3단 배치 시 빈 열을 기준으로 블록 분리."""
    if raw is None or raw.empty:
        return []

    cols_nonempty = [int(raw.iloc[:, i].notna().sum()) for i in range(raw.shape[1])]
    is_empty_col = [cnt == 0 for cnt in cols_nonempty]

    blocks: List[tuple[int, int]] = []
    start = None
    for i, empty in enumerate(is_empty_col):
        if not empty and start is None:
            start = i
        elif empty and start is not None:
            if i - start >= min_nonempty_cols:
                blocks.append((start, i))
            start = None
    if start is not None and raw.shape[1] - start >= min_nonempty_cols:
        blocks.append((start, raw.shape[1]))

    # 블록 감지 실패 시 전체 1블록
    if not blocks:
        blocks = [(0, raw.shape[1])]

    out = []
    for s, e in blocks:
        b = raw.iloc[:, s:e].copy()
        b = b.dropna(how="all").dropna(axis=1, how="all")
        if not b.empty:
            out.append(b.reset_index(drop=True))
    return out


# ---------------------------------------------------------------------------
# 헤더 탐지 및 구조 복원
# ---------------------------------------------------------------------------
def _find_header_row(block: pd.DataFrame, max_scan: int = 20) -> Optional[int]:
    scan_n = min(max_scan, len(block))
    best_idx = None
    best_score = -1
    for i in range(scan_n):
        row = block.iloc[i].tolist()
        score = 0
        if any(_contains_any(x, NAME_KEYS) for x in row):
            score += 3
        if any(_contains_any(x, DEPT_KEYS) for x in row):
            score += 2
        if any(_contains_any(x, TITLE_KEYS) for x in row):
            score += 1
        if any(_contains_any(x, PHONE_KEYS) for x in row):
            score += 1
        if any(_contains_any(x, EMAIL_KEYS) for x in row):
            score += 1
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score > 0 else None


def _make_unique_columns(row: List[object]) -> List[str]:
    cols = []
    seen = {}
    for i, v in enumerate(row):
        name = _clean_text(v) or f"col_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        cols.append(name)
    return cols


def _pick_col(cols: List[str], keys: Iterable[str]) -> Optional[str]:
    for c in cols:
        if _contains_any(c, keys):
            return c
    return None


def _rowwise_org_fill(df: pd.DataFrame) -> pd.DataFrame:
    """병합셀 해제 후에도 남는 조직 단서를 좌→우, 상→하로 보강."""
    if df is None or df.empty:
        return pd.DataFrame()
    d = df.copy()
    d = d.ffill(axis=0)
    d = d.ffill(axis=1)
    return d


def _extract_rows_from_block(block: pd.DataFrame, sheet_name: str, block_index: int) -> ParseBlockResult:
    header_idx = _find_header_row(block)

        # 1) 정형 테이블인 경우
    if header_idx is not None:
        work = block.copy()
        work = _rowwise_org_fill(work)

        cols = _make_unique_columns(work.iloc[header_idx].tolist())
        body = work.iloc[header_idx + 1 :].copy().reset_index(drop=True)
        body.columns = cols
        body = body.dropna(how="all")

        rows = []

        # 성명/이름 컬럼이 나오는 위치를 모두 찾는다
        name_positions = [i for i, c in enumerate(cols) if _contains_any(c, NAME_KEYS)]

        if not name_positions:
            return ParseBlockResult(sheet_name, block_index, pd.DataFrame())

        for _, r in body.iterrows():
            for name_idx in name_positions:
                vals = r.tolist()

                name = _clean_text(vals[name_idx]) if name_idx < len(vals) else ""
                if not name:
                    continue

                # 이름 앞쪽 1~4칸에서 조직 후보 찾기
                org_candidates = []
                for j in range(max(0, name_idx - 4), name_idx):
                    vv = _clean_text(vals[j])
                    if vv:
                        org_candidates.append(vv)

                dept_raw = ""
                if org_candidates:
                    dept_raw = " / ".join(org_candidates)

                # 이름 오른쪽 1~4칸에서 직책/전화/메일 찾기
                title = ""
                phone = ""
                email = ""
                ext = ""

                for j in range(name_idx + 1, min(len(vals), name_idx + 5)):
                    vv = _clean_text(vals[j])
                    if not vv:
                        continue
                    if _contains_any(vv, TITLE_KEYS) and not title:
                        title = vv
                    elif _looks_like_phone(vv) and not phone:
                        phone = vv
                    elif _looks_like_email(vv) and not email:
                        email = vv
                    elif _contains_any(vv, EXT_KEYS) and not ext:
                        ext = vv

                org_levels = _split_org_path_from_text(dept_raw)

                rows.append({
                    "org_level_1": org_levels[0],
                    "org_level_2": org_levels[1],
                    "org_level_3": org_levels[2],
                    "org": org_levels[0] or org_levels[1] or org_levels[2] or dept_raw,
                    "dept": org_levels[2] or org_levels[1] or dept_raw,
                    "name": name,
                    "title": title,
                    "phone": phone or ext,
                    "email": email,
                    "source_sheet": sheet_name,
                    "source_block": block_index,
                    "parse_mode": "tabular_contact_multi",
                })

        return ParseBlockResult(sheet_name, block_index, pd.DataFrame(rows))

    # 2) 비정형 연락망인 경우: 행 안의 조직명/이름/연락처를 휴리스틱으로 추출
    work = _rowwise_org_fill(block)
    rows = []
    current_l1 = ""
    current_l2 = ""
    current_l3 = ""

    for _, rr in work.iterrows():
        vals = [_clean_text(x) for x in rr.tolist()]

        for i, v in enumerate(vals):

            # 🔥 조직 먼저 업데이트
            if _looks_like_org_name(v):
                current_l1, current_l2, current_l3 = _merge_org_context(
                    current_l1, current_l2, current_l3, [v]
                )
                continue

            # 🔥 이름 찾기
            if len(v) >= 2 and len(v) <= 20:

                name = v

                title = ""
                phone = ""
                email = ""

                for j in range(max(0, i-2), min(len(vals), i+3)):
                    vv = vals[j]

                    if _looks_like_phone(vv):
                        phone = vv
                    elif _looks_like_email(vv):
                        email = vv
                    elif _contains_any(vv, TITLE_KEYS):
                        title = vv

                rows.append({
                    "org_level_1": current_l1,
                    "org_level_2": current_l2,
                    "org_level_3": current_l3,
                    "org": current_l1 or current_l2 or current_l3,
                    "dept": current_l3 or current_l2,
                    "name": name,
                    "title": title,
                    "phone": phone,
                    "email": email,
                    "source_sheet": sheet_name,
                    "source_block": block_index,
                    "parse_mode": "cell_scan",
                })
        

    return ParseBlockResult(sheet_name, block_index, pd.DataFrame(rows))


def _split_org_path_from_text(text: str) -> tuple[str, str, str]:
    t = _clean_text(text)
    if not t:
        return "", "", ""
    parts = [p.strip() for p in t.replace(">", "/").replace("|", "/").split("/") if p.strip()]
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return parts[0], parts[1], parts[1]
    # 단일 값이면 성격 추정
    if any(k in t for k in ["본부", "사업부", "사업본부"]):
        return t, "", ""
    if any(k in t for k in ["그룹", "센터", "실"]):
        return "", t, ""
    return "", "", t


def _merge_org_context(cur1: str, cur2: str, cur3: str, org_candidates: List[str]) -> tuple[str, str, str]:
    l1, l2, l3 = cur1, cur2, cur3
    for org in org_candidates:
        o = _clean_text(org)
        if any(k in o for k in ["본부", "사업부", "사업본부"]):
            l1 = o
            l2 = ""
            l3 = ""
        elif any(k in o for k in ["그룹", "센터", "실"]):
            l2 = o
            l3 = ""
        elif any(k in o for k in ["팀", "파트", "lab", "랩"]):
            l3 = o
    return l1, l2, l3


# ---------------------------------------------------------------------------
# 외부 호출 함수
# ---------------------------------------------------------------------------
def parse_contact_org_xlsx(path: str | Path, sheet_names: Optional[List[str]] = None) -> pd.DataFrame:
    """연락망형 Excel을 읽어 조직 계층 + 인물 표준 DataFrame 반환."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)

    wb = load_workbook(p, data_only=True)
    targets = sheet_names or wb.sheetnames
    results = []

    for sh in targets:
        ws = wb[sh]
        raw = _worksheet_to_df_merged_values(ws)
        if raw.empty:
            continue
        blocks = _split_wide_blocks(raw)
        for bi, block in enumerate(blocks, start=1):
            parsed = _extract_rows_from_block(block, sh, bi).rows
            if parsed is not None and not parsed.empty:
                parsed["source_file"] = p.name
                results.append(parsed)

    if not results:
        return pd.DataFrame(
            columns=[
                "org_level_1", "org_level_2", "org_level_3",
                "org", "dept", "name", "title", "phone", "email",
                "source_file", "source_sheet", "source_block", "parse_mode",
            ]
        )

    out = pd.concat(results, ignore_index=True)
    out = out.drop_duplicates(subset=["org_level_1", "org_level_2", "org_level_3", "name", "title", "phone", "email"])
    return out.reset_index(drop=True)


def save_contact_org_csv(path: str | Path, out_csv: str | Path) -> pd.DataFrame:
    df = parse_contact_org_xlsx(path)
    out = Path(out_csv)
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out, index=False, encoding="utf-8-sig")
    return df


if __name__ == "__main__":
    sample = Path("data/raw/아이엔소프트_직원연락망 (23.06.01).xlsx")
    if sample.exists():
        df = parse_contact_org_xlsx(sample)
        print(df.head(20).to_string())
        print(f"rows={len(df)}")
    else:
        print("sample file not found")


# ===== 연락망 전용 파서 =====

def clean_merged_cells(df):
    return df.fillna(method="ffill")


def split_wide_table(df):
    chunks = []

    cols = df.columns.tolist()
    step = 5

    for i in range(0, len(cols), step):
        sub_cols = cols[i:i+step]

        if len(sub_cols) < 3:
            continue

        sub_df = df[sub_cols].copy()

        try:
            sub_df.columns = ["org", "name", "title", "phone", "email"][:len(sub_cols)]
        except:
            continue

        chunks.append(sub_df)

    if chunks:
        return pd.concat(chunks, ignore_index=True)

    return pd.DataFrame()


def parse_contact_sheet(df):

    df = df.ffill()

    # ===== 컬럼 rename 먼저 =====
    rename_map = {
        "성명": "name",
        "이름": "name",
        "사원명": "name",
        "부서": "org",
        "부서.1": "org",
        "직책": "title",
        "휴대폰": "phone",
        "e-mail": "email",
        "e-mail.1": "email"
    }

    df = df.rename(columns=rename_map)

    # 🔥 중복 컬럼 제거 (핵심)
    df = df.loc[:, ~df.columns.duplicated()]

    # ===== name 컬럼 없으면 종료 =====
    if "name" not in df.columns:
        return pd.DataFrame()

    # ===== 필요한 컬럼만 선택 =====
    cols = [c for c in ["name", "org", "title", "phone", "email"] if c in df.columns]

    df = df[cols]

    # ===== name 기준 필터 =====
    df = df[df["name"].notna()]

    return df
